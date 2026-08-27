"""The assumption layer: analyst inputs with provenance, loaded, saved and seeded.

The rules under test are the roadmap's: every number is an analyst input carrying its
source, a scenario restates only what it changes, and the seed file reconstructs the
whole layer on a database rebuilt from nothing.
"""

import db
import assumptions as A


def _seed(tmp_path):
    path = str(tmp_path / "a.db")
    db.init(path)
    conn = db.get_connection(path)
    conn.execute("INSERT INTO companies (id, ticker, name) VALUES (1, 'VRTX', 'Vertex')")
    conn.execute("INSERT INTO assets (id, owner_company_id, brand_name, generic_name,"
                 " modality, is_marketed) VALUES"
                 " (1, 1, 'Casgevy', 'Exagamglogene autotemcel', 'biologic', 1)")
    conn.execute("INSERT INTO indications (id, name) VALUES (1, 'Anemia, Sickle Cell')")
    conn.commit()
    return path, conn


def test_save_and_load_round_trip(tmp_path):
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [
        {"key": "therapy_mode", "text_value": "one_time", "source": "workbook"},
        {"key": "net_price_per_patient", "value": 1.8, "unit": "mm USD",
         "source": "Vertex Dec 2023"},
        {"key": "prevalence", "value": 100000, "indication_id": 1, "source": "CDC"},
        {"key": "new_patients", "value": 5, "year": 2024, "indication_id": 1,
         "source": "analyst"},
        {"key": "new_patients", "value": 45, "year": 2025, "indication_id": 1,
         "source": "analyst"},
    ])
    conn.commit()
    got = A.load(conn, 1)
    assert got["scalars"]["therapy_mode"] == "one_time"
    assert got["scalars"]["net_price_per_patient"] == 1.8
    ind = got["indications"][0]
    assert ind["name"] == "Anemia, Sickle Cell"
    assert ind["scalars"]["prevalence"] == 100000
    assert ind["series"]["new_patients"] == {2024: 5, 2025: 45}
    conn.close()


def test_writing_the_same_key_replaces_rather_than_duplicates(tmp_path):
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [{"key": "net_price_per_patient", "value": 1.8, "source": "a"}])
    A.save(conn, 1, [{"key": "net_price_per_patient", "value": 1.65, "source": "b"}])
    conn.commit()
    rows = A.rows(conn, 1)
    assert len(rows) == 1
    assert rows[0]["value"] == 1.65 and rows[0]["source"] == "b"
    conn.close()


def test_an_empty_value_deletes_the_row(tmp_path):
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [{"key": "sga_pct", "value": 0.2}])
    A.save(conn, 1, [{"key": "sga_pct", "value": None}])
    conn.commit()
    assert A.rows(conn, 1) == []
    conn.close()


def test_a_row_with_no_key_is_refused(tmp_path):
    path, conn = _seed(tmp_path)
    try:
        A.save(conn, 1, [{"value": 1.0}])
        raised = False
    except ValueError:
        raised = True
    assert raised
    conn.close()


def test_a_scenario_restates_only_what_it_changes(tmp_path):
    """A bear case that touches three numbers must not need the other forty copied."""
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [
        {"key": "net_price_per_patient", "value": 1.8, "source": "base"},
        {"key": "sga_pct", "value": 0.2, "source": "base"},
        {"key": "net_price_per_patient", "value": 1.5, "scenario": "bear",
         "source": "bear case"},
    ])
    conn.commit()
    bear = A.load(conn, 1, scenario="bear")
    assert bear["scalars"]["net_price_per_patient"] == 1.5   # restated
    assert bear["scalars"]["sga_pct"] == 0.2                 # inherited from base
    conn.close()


def test_the_seed_file_rebuilds_the_layer_from_nothing(tmp_path):
    path, conn = _seed(tmp_path)
    seed_dir = tmp_path / "assumptions"
    seed_dir.mkdir()
    (seed_dir / "x.csv").write_text(
        "# comment\n"
        "ticker,brand,indication,region,scenario,key,year,value,text_value,unit,"
        "source,note\n"
        "VRTX,Casgevy,,US,base,therapy_mode,,,one_time,,workbook,\n"
        'VRTX,Casgevy,"Anemia, Sickle Cell",US,base,prevalence,,100000,,patients,CDC,\n'
        "VRTX,Casgevy,\"Anemia, Sickle Cell\",US,base,new_patients,2024,5,,patients,"
        "analyst,\n"
        "OTHR,Missing,,US,base,therapy_mode,,,one_time,,x,\n")
    out = A.load_seeds(conn, seed_dir)
    assert out["written"] == 3
    assert out["skipped"] == 1                      # the unknown product is skipped
    got = A.load(conn, 1)
    assert got["scalars"]["therapy_mode"] == "one_time"
    assert got["indications"][0]["series"]["new_patients"] == {2024: 5}
    conn.close()


def test_seeding_twice_is_idempotent(tmp_path):
    path, conn = _seed(tmp_path)
    seed_dir = tmp_path / "assumptions"
    seed_dir.mkdir()
    (seed_dir / "x.csv").write_text(
        "ticker,brand,indication,region,scenario,key,year,value,text_value,unit,"
        "source,note\n"
        "VRTX,Casgevy,,US,base,sga_pct,,0.2,,,workbook,\n")
    A.load_seeds(conn, seed_dir)
    A.load_seeds(conn, seed_dir)
    assert len(A.rows(conn, 1)) == 1
    conn.close()


def test_the_shipped_casgevy_seed_parses_and_carries_sources():
    """The real file: every row a source or an explicit for-review note, so a typo fails
    here rather than silently seeding nothing."""
    import csv
    path = A.SEED_DIR / "casgevy.csv"
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(
            [l for l in handle if not l.lstrip().startswith("#")]))
    assert len(rows) > 40
    for row in rows:
        assert row["ticker"] == "VRTX" and row["brand"] == "Casgevy"
        assert (row["source"] or "").strip() or "for review" in (row["note"] or ""), \
            f"unsourced row without a for-review note: {row['key']}"


def test_the_default_files_parse_and_carry_sources():
    pos = A.pos_defaults()
    assert "Phase 3" in pos and pos["Phase 3"]["pos"] == 0.65
    assert all(row["source"] for row in pos.values())
    erosion = A.erosion_defaults()
    assert "small molecule" in erosion
    assert erosion["small molecule"]["year1_pct"] > erosion["biologic"]["year1_pct"]
    assert all(row["source"] for row in erosion.values())


def test_snapshot_writes_forecast_history(tmp_path):
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [{"key": "sga_pct", "value": 0.2, "source": "x"}])
    A.snapshot(conn, 1, "base", {"rnpv": 1911.7, "npv": 2367.5, "wacc": 0.098,
                                 "pos": 0.8075, "revenue_after_loe": [1.0],
                                 "years": [2026]})
    conn.commit()
    row = conn.execute("SELECT source, entity_key, payload FROM snapshots"
                       " WHERE source = 'forecast'").fetchone()
    assert row["entity_key"] == "1"
    import json
    payload = json.loads(row["payload"])
    assert payload["rnpv"] == 1911.7
    assert payload["assumptions"][0]["key"] == "sga_pct"
    conn.close()


def test_xlsx_export_round_trips_through_openpyxl(tmp_path):
    import io

    from openpyxl import load_workbook
    path, conn = _seed(tmp_path)
    A.save(conn, 1, [{"key": "net_price_per_patient", "value": 1.8,
                      "source": "Vertex"}])
    blob = A.export_xlsx(conn, 1, "base", None)
    book = load_workbook(io.BytesIO(blob))
    sheet = book["Assumptions"]
    header = [c.value for c in sheet[1]]
    assert "key" in header and "source" in header
    values = {row[header.index("key")]: row[header.index("value")]
              for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert values["net_price_per_patient"] == 1.8
    conn.close()


def test_a_seed_can_name_a_compound_that_has_no_brand_yet(tmp_path):
    """Retatrutide, milvexian and every other Phase 3 asset carries a generic name and a
    null brand. A loader matching brand alone could seed marketed products and nothing in
    the pipeline, which is where a forecast is worth most."""
    path, conn = _seed(tmp_path)
    conn.execute("INSERT INTO assets (id, owner_company_id, generic_name, is_marketed)"
                 " VALUES (2, 1, 'Retatrutide', 0)")
    conn.commit()
    directory = tmp_path / "seeds"
    directory.mkdir()
    (directory / "reta.csv").write_text(
        "ticker,brand,indication,region,scenario,key,year,value,text_value,unit,source,note\n"
        "VRTX,Retatrutide,,US,base,tax_rate,,0.2,,,10-K,\n", encoding="utf-8")

    assert A.load_seeds(conn, directory)["written"] == 1
    got = A.load(conn, 2)
    assert got["scalars"]["tax_rate"] == 0.2
    conn.close()


def test_a_brand_still_wins_over_a_compound_of_the_same_name(tmp_path):
    path, conn = _seed(tmp_path)
    # A launched product must never be shadowed by a compound sharing its ingredient.
    conn.execute("INSERT INTO assets (id, owner_company_id, generic_name, is_marketed)"
                 " VALUES (3, 1, 'Casgevy', 0)")
    conn.commit()
    directory = tmp_path / "seeds"
    directory.mkdir()
    (directory / "c.csv").write_text(
        "ticker,brand,indication,region,scenario,key,year,value,text_value,unit,source,note\n"
        "VRTX,Casgevy,,US,base,tax_rate,,0.15,,,workbook,\n", encoding="utf-8")
    A.load_seeds(conn, directory)

    assert A.load(conn, 1)["scalars"].get("tax_rate") == 0.15   # the marketed row
    assert not A.load(conn, 3)["scalars"]
    conn.close()
